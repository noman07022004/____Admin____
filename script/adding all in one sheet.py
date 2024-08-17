'''
Create a Python script that processes an Excel file containing multiple sheets named sequentially as "chapter 1," "chapter 2," etc., combines the non-empty rows from these sheets into a new sheet, and saves this combined data into a new Excel file. The new file should be named by appending "(all in one sheet)" to the original file name. The script should also print a detailed summary of the combined data and handle any errors gracefully.
'''

from openpyxl import load_workbook, Workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from pprint import pprint
import os

def select_excel_file():
    # Hide the main Tkinter window
    Tk().withdraw()
    filename = askopenfilename(
        title="Select an Excel file",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    return filename

def is_not_empty(value):
    """ Helper function to check if a cell value is not empty. """
    return value is not None and value != ""

def generate_new_filename(original_filename):
    # Generate new filename by adding "(all in one sheet)" before the extension
    base, ext = os.path.splitext(original_filename)
    new_filename = f"{base} (all in one sheet){ext}"
    return new_filename

def combine_sheets(excel_file):
    try:
        # Load the original workbook
        workbook = load_workbook(excel_file)

        # List all sheet names
        sheet_names = workbook.sheetnames

        # Create a new workbook for combined data
        new_workbook = Workbook()
        new_sheet = new_workbook.active
        new_sheet.title = "all in one"
        
        # Combine data from chapter sheets into the new sheet
        combined_data = []
        combined_order = []

        for sheet_name in sorted(sheet_names):
            if sheet_name.lower().startswith("chapter"):
                sheet = workbook[sheet_name]
                non_empty_row_count = 0
                for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
                    cell_value = row[0]
                    if is_not_empty(cell_value):
                        combined_data.append(cell_value)
                        non_empty_row_count += 1

                # Append sheet name with non-empty row count
                combined_order.append((sheet_name, non_empty_row_count))

        # Write combined data to the new sheet, skipping empty values
        for index, value in enumerate(combined_data, start=1):
            new_sheet.cell(row=index, column=1, value=value)

        # Generate new file name
        new_filename = generate_new_filename(excel_file)
        
        # Save the new workbook
        new_workbook.save(new_filename)
        
        # Print the combined order in a beautiful way
        print("\nCombined Order of Sheets:")
        print("==========================")
        for idx, (name, row_count) in enumerate(combined_order, start=1):
            print(f"{idx}. {name} - {row_count} non-empty rows")
        print("==========================\n")
        
        # Summary of the combination
        pprint(f"Combined data from {len(combined_order)} sheets into 'all in one'.")
        pprint(f"Total non-empty rows combined: {len(combined_data)}")
        pprint(f"New file created: {new_filename}")

    except Exception as e:
        pprint(f"An error occurred: {e}")

if __name__ == "__main__":
    # Ask the user to select an Excel file
    excel_file = select_excel_file()

    # Check if a file was selected
    if excel_file:
        combine_sheets(excel_file)
    else:
        pprint("No file selected.")
